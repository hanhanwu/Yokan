import pickle
import os
import io
import re
from googleapiclient.discovery import build
from googleapiclient.http import MediaInMemoryUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl.styles import Alignment


FINANCIAL_RAG_SYSTEM_PROMPT = """You are a finance expert.
Your role is to answer financial questions with precision and clarity.

GUIDELINES:
- Answer only use retrieved content as reference, do NOT use any other information
- If the retrieved content does not contain relevant information to answer the question, say "I don't know" instead of making up an answer
- If data is missing or unclear, state it explicitly - do NOT make up numbers
- Include relevant financial metrics and ratios in your analysis
- Flag any assumptions you make about the data
- For complex queries, structure responses with clear breakdowns

FINANCIAL ACCURACY IS CRITICAL. When in doubt, cite your source and indicate uncertainty.
"""


def json_default(obj):
    if isinstance(obj, set):
        return list(obj)
    raise TypeError(f'Object of type {obj.__class__.__name__} is not JSON serializable')


def upload_to_google_drive(df, folder_id, output_filename):
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    GCP_DIR = os.environ.get("GCP_CREDENTIALS_DIR", os.path.join(os.path.expanduser("~"), ".gcp"))
    TOKEN_FILE = os.path.join(GCP_DIR, "token.pickle")
    CREDENTIALS_FILE = os.path.join(GCP_DIR, "drive_auth.json")

    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as f:
            creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as f:
            pickle.dump(creds, f)

    service = build('drive', 'v3', credentials=creds)

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')

    from openpyxl import load_workbook
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    media = MediaInMemoryUpload(buffer.read(), mimetype=mimetype)

    # check if file already exists in the folder
    query = f"name='{output_filename}' and '{folder_id}' in parents and trashed=false"
    existing = service.files().list(q=query, fields='files(id)').execute().get('files', [])

    if existing:
        file_id = existing[0]['id']
        uploaded = service.files().update(fileId=file_id, media_body=media, fields='id').execute()
        print(f"Overwritten file ID: {uploaded.get('id')}")
    else:
        file_metadata = {'name': output_filename, 'parents': [folder_id]}
        uploaded = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"Uploaded file ID: {uploaded.get('id')}")


# ============================================================================
# TEST NEW AUTO EVALUATION
# ============================================================================
from pydantic import BaseModel, Field
import pandas as pd
import yaml
import os
import json
import asyncio
import psycopg2
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from groq import RateLimitError

from langchain_groq import ChatGroq
from langchain.prompts import PromptTemplate
from langchain.output_parsers import PydanticOutputParser
from langchain.output_parsers import OutputFixingParser

_here = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(_here, 'eval_prompts.yaml'), 'r') as file:
    prompt_versions = yaml.safe_load(file)

eval_llm = ChatGroq(
    groq_api_key=os.environ["GROQ_TOKEN"],
    model_name="openai/gpt-oss-20b", 
    temperature=0.7
)

@retry(
    retry=retry_if_exception_type(RateLimitError),
    wait=wait_exponential(multiplier=2, min=5, max=60),
    stop=stop_after_attempt(5)
)
async def _invoke_with_retry(chain, inputs):
    return await chain.ainvoke(inputs)


def get_eval_input(db_url, config_hash):
    conn = psycopg2.connect(
        host=db_url.host,
        port=db_url.port,
        dbname=db_url.database,
        user=db_url.username,
        password=db_url.password,
    )
    cur = conn.cursor()
    cur.execute("SELECT output FROM existing_rag_output WHERE config_hash = %s", 
                (config_hash,))
    row = cur.fetchone()
    json_results = row[0]
    cur.close()
    conn.close()

    records = []
    for item in json_results:
        record = {
            'query': item['question'],
            'ai_answer': item['ai_answer'],
            'referenced_answer': item['expected_answer'],
            'retrieved_content': ''.join(content_dct['content'] for content_dct in item['retrieved_lst']),
        }
        records.append(record)
    return pd.DataFrame(records)


async def eval_one_config(config_hash, db_url, rag_df):    
    input_df = get_eval_input(db_url, config_hash)
    input_df = pd.merge(input_df, rag_df[['question', 'context']], 
                        left_on='query', right_on='question')
    input_df.drop(columns=['question'], inplace=True)

    retrieval_quality = await get_retrieval_quality_output_async(input_df, eval_llm,
                                                            prompt_versions['rq_prompt_template'])
    retrieval_quality['same_context'] = retrieval_quality['retrieved_content'] == retrieval_quality['context']
    
    answer_quality = await get_answer_quality_output_async(input_df, eval_llm,
                                                            prompt_versions['aq_prompt_template'])

    return config_hash, retrieval_quality, answer_quality


async def run_auto_eval(config_hashes, db_url, rag_df, max_concurrent_configs=1):
    sem = asyncio.Semaphore(max_concurrent_configs)  # Semaphore to limit concurrent runs of different configs

    async def throttled_eval(config_hash):
        async with sem:
            return await eval_one_config(config_hash, db_url, rag_df)

    results = await asyncio.gather(*[
        throttled_eval(config_hash)
        for config_hash in config_hashes
    ])

    return {
         config_hash: {"retrieval_quality_df": retrieval_quality,
                        "answer_quality_df": answer_quality}
         for config_hash, retrieval_quality, answer_quality in results
     }


# ------------------------------------------ RETRIEVAL QUALITY ------------------------------------------ #
class RetrievalQuality(BaseModel):
    score: int = Field(description="""Score with:
                - Only generate the score as -1, 0 or 1 or 2 or 3
                - Scoring as -1: if the RETRIEVED CONTENT is much more relevant to the USER QUERY than the CONTEXT
                - Scoring as 0: if the RETRIEVED CONTENT is completely irrelevant to the USER QUERY
                - If the CONTEXT is strongly relevant to the USER QUERY:
                    - Scoring as 1: if the RETRIEVED CONTENT is relevant to the USER QUERY but doesn't contain any critical information from the CONTEXT
                    - Scoring as 2: if the RETRIEVED CONTENT is relevant to the USER QUERY but only partially contains critical information from the CONTEXT
                    - Scoring as 3: if the RETRIEVED CONTENT is relevant to USER QUERY and contains all the critical information from the CONTEXT
            """)
    reasoning: str = Field(description="Reasoning for the given score.")


async def evaluate_retrieval_quality_async(llm, user_query, context, retrieved_content, rq_prompt_template):
    base_parser = PydanticOutputParser(pydantic_object=RetrievalQuality)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=rq_prompt_template,
        input_variables=["user_query", "context", "retrieved_content"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query,
        "context": context,
        "retrieved_content": retrieved_content
    })
    return result


async def process_retrieval_quality_record_async(llm, record, rq_prompt_template):
    eval_result = await evaluate_retrieval_quality_async(
        llm,
        record['query'],
        record['context'],
        record['retrieved_content'],
        rq_prompt_template
    )
    record['retrieval_quality_score'] = eval_result.score
    record['rq_reasoning'] = eval_result.reasoning
    return record


async def get_retrieval_quality_output_async(input_df, llm, rq_prompt_template, concurrency=2):
    sem = asyncio.Semaphore(concurrency)  # Semaphore to throttle concurrency for running records in parallel

    async def sem_task(record):
        async with sem:
            return await process_retrieval_quality_record_async(llm, record, rq_prompt_template)

    input_records = input_df.to_dict(orient='records')
    tasks = [sem_task(record) for record in input_records]
    output_lst = await asyncio.gather(*tasks)
    output_df = pd.DataFrame(output_lst)
    return output_df
# ------------------------------------ QUERY QUALITY ------------------------------------ #


# ------------------------------------ ANSWER QUALITY ------------------------------------ #
class AnswerQuality(BaseModel):
    score: int = Field(description="""Score with:
    - Only generate the score as -1, 0 or 1 or 2 or 3
    - Scoring as -1: if the AI's ANSWER is much more relevant to the USER QUERY than the REFERENCED ANSWER
    - Scoring as 0: if the AI's ANSWER is completely irrelevant to the USER QUERY
    - If the REFERENCED ANSWER is strongly relevant to the USER QUERY:
        - Scoring as 0: if AI's ANSWER is significantly conflict with the REFERENCED ANSWER
        - Scoring as 1: if the AI's ANSWER is relevant to the USER QUERY but doesn't contain any critical information from the REFERENCED ANSWER
        - Scoring as 2: if the AI's ANSWER is relevant to the USER QUERY but only partially contains critical information from the REFERENCED ANSWER
        - Scoring as 3: if the AI's ANSWER is relevant to USER QUERY and contains all the critical information from the REFERENCED ANSWER
    """)
    reasoning: str = Field(description="Reasoning for the given score.")


async def evaluate_answer_quality_async(llm, user_query, ai_answer, referenced_answer, aq_prompt_template):
    base_parser = PydanticOutputParser(pydantic_object=AnswerQuality)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=aq_prompt_template,
        input_variables=["user_query", "ai_answer", "referenced_answer"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query,
        "ai_answer": ai_answer,
        "referenced_answer": referenced_answer
    })
    return result


async def process_answer_quality_record_async(llm, record, aq_prompt_template):
    eval_result = await evaluate_answer_quality_async(
        llm,
        record['query'],
        record['ai_answer'],
        record['referenced_answer'],
        aq_prompt_template
    )
    record['answer_quality_score'] = eval_result.score
    record['aq_reasoning'] = eval_result.reasoning
    return record


async def get_answer_quality_output_async(input_df, llm, aq_prompt_template, concurrency=2):
    sem = asyncio.Semaphore(concurrency)  # Semaphore to throttle concurrency for running records in parallel

    async def sem_task(record):
        async with sem:
            return await process_answer_quality_record_async(llm, record, aq_prompt_template)

    input_records = input_df.to_dict(orient='records')
    tasks = [sem_task(record) for record in input_records]
    output_lst = await asyncio.gather(*tasks)
    output_df = pd.DataFrame(output_lst)
    return output_df
# ------------------------------------ ANSWER QUALITY ------------------------------------ #


# ============================================================================
# ROOT CAUSE ANALYSIS AGENT SYSTEM
# ============================================================================
def contains_conflict(text: str) -> bool:
    pattern = r'\b(conflict\w*|contradict\w*)\b'
    return re.search(pattern, text, re.IGNORECASE) is not None


def get_auto_eval_output(db_url):
    conn = psycopg2.connect(
        host=db_url.host,
        port=db_url.port,
        dbname=db_url.database,
        user=db_url.username,
        password=db_url.password,
    )
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t1.config_hash,
            t1.dataset,
            t1.embedding_model,
            t1.top_n_retrieval,
            t1.semantic_weight,
            t1.answer_gen_llm,
            rq.value->>'query' AS query,
            rq.value->>'context' AS context,
            rq.value->>'retrieved_content' AS retrieved_content,
            rq.value->>'same_context' AS same_context,
            rq.value->>'retrieval_quality_score' AS retrieval_quality_score,
            rq.value->>'rq_reasoning' AS rq_reasoning,
            aq.value->>'referenced_answer' AS referenced_answer,
            aq.value->>'ai_answer' AS ai_answer,
            aq.value->>'answer_quality_score' AS answer_quality_score,
            aq.value->>'aq_reasoning' AS aq_reasoning
        FROM existing_rag_output AS t1
        JOIN existing_auto_eval_output AS t2
            ON t1.config_hash = t2.config_hash,
        jsonb_array_elements(t2.retrieval_quality) WITH ORDINALITY AS rq(value, idx),
        jsonb_array_elements(t2.answer_quality) WITH ORDINALITY AS aq(value, idx)
        WHERE rq.idx = aq.idx
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    col_names = [desc[0] for desc in cur.description]
    df = pd.DataFrame(rows, columns=col_names)

    return df


# ------------------------------------------ SCORE AFTER REVIEW ------------------------------------------ #
class ScoreAfterReview(BaseModel):
    score: int = Field(description="Only generate an integer score in [-1, 0, 1, 2, 3]")


async def review_sr_async(llm, eval_reasoning):
    base_parser = PydanticOutputParser(pydantic_object=ScoreAfterReview)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=prompt_versions['review_reasoning_prompt_template'],
        input_variables=["eval_reasoning"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "eval_reasoning": eval_reasoning
    })
    return result.score
# ------------------------------------------ SCORE AFTER REVIEW ------------------------------------------ #


# ------------------------------------------ TEXT ALIGNMENT ------------------------------------------ #
class TextAlignment(BaseModel):
    score: int = Field(description="Only generate an integer score as 0 or 1. 1 means aligned, 0 means not aligned.")
    reasoning: str = Field(description="Reasoning for the given score.")


async def review_ta_async(llm, user_query, text_a, text_b):
    base_parser = PydanticOutputParser(pydantic_object=TextAlignment)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=prompt_versions['review_text_alignment_template'],
        input_variables=["user_query", "text_a", "text_b"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query,
        "text_a": text_a,
        "text_b": text_b
    })
    return result
# ------------------------------------------ TEXT ALIGNMENT ------------------------------------------ #


# ------------------------------------------ QUERY QUALITY ------------------------------------------ #
class QueryQuality(BaseModel):
    query_quality: str = Field(description="Only generate a value from ['clear', 'ambiguous']")


async def review_query_quality_async(llm, user_query):
    base_parser = PydanticOutputParser(pydantic_object=QueryQuality)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=prompt_versions['review_query_quality_template'],
        input_variables=["user_query"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query
    })
    return result.query_quality
# ------------------------------------------ QUERY QUALITY ------------------------------------------ #


# ------------------------------------------ QUERY EXPANSION ------------------------------------------ #
class QueryExpansion(BaseModel):
    query_variants: list[str] = Field(description="A list of 1 to 3 clearer, specific query variants.")


async def expand_query_async(llm, user_query):
    base_parser = PydanticOutputParser(pydantic_object=QueryExpansion)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=prompt_versions['query_expansion_template'],
        input_variables=["user_query"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query
    })
    return result.query_variants
# ------------------------------------------ QUERY EXPANSION ------------------------------------------ #


# ------------------------------------------ REVIEW RAG SYSTEM ------------------------------------------ #
class RAGSystemReview(BaseModel):
    root_cause_analysis: str = Field(description="Explain potential root causes of RAG's answer quality scores.")
    improvement_suggestions: str = Field(description="Provide suggestions to improve the RAG performance.")


async def review_rag_system_async(llm, avg_rq_score, rq_reasons, avg_aq_score, aq_reasons,
                                system_prompt, rag_config):
    base_parser = PydanticOutputParser(pydantic_object=RAGSystemReview)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=prompt_versions['review_rag_system_template'],
        input_variables=["avg_rq_score", "rq_reasons", "avg_aq_score", "aq_reasons",
                        "system_prompt", "rag_config"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "avg_rq_score": avg_rq_score, "rq_reasons": rq_reasons,
        "avg_aq_score": avg_aq_score, "aq_reasons": aq_reasons,
        "system_prompt": system_prompt, "rag_config": rag_config
    })
    return result
# ------------------------------------------ REVIEW RAG SYSTEM ------------------------------------------ #